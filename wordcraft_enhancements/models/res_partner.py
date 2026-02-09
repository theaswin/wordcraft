from odoo import models, fields, _ , api
from odoo.exceptions import UserError
import base64
from io import BytesIO
from openpyxl import load_workbook
import base64
import csv
from io import StringIO
import openpyxl
import io
class ResPartner(models.Model):
    _inherit = "res.partner"

    mobile = fields.Char(string="Mobile")
    google_map_url = fields.Char(string="Google Map")

    category_import_file = fields.Binary(
        string="Import Categories (Excel)",
        help="XLSX columns: name, mobile, category_name"
    )
    category_import_filename = fields.Char(string="Filename")
    due_amount = fields.Float(string="Amount Due")
    payment_amount_due = fields.Float(string="Payment Amount Due")
    payment_amount_due_amt = fields.Float(string="Payment Amount Due (Amount)")
    payment_amount_due_amt_supplier = fields.Float(string="Payment Amount Due (Amount)")
    payment_amount_overdue = fields.Float(string="Overdue")


    due_invoice_count = fields.Integer(
        string="Due Invoice Count",
        compute="_compute_due_invoice_count",
        store=False
    )

    @api.depends('invoice_ids')
    def _compute_due_invoice_count(self):
        Move = self.env['account.move']
        for rec in self:
            rec.due_invoice_count = Move.search_count([
                ('partner_id', '=', rec.id),
                ('move_type', '=', 'out_invoice'),
                ('state', '=', 'posted'),
                ('amount_residual', '>', 0),
            ])
    

    def action_import_partner_categories(self):
        self.ensure_one()

        if not self.category_import_file:
            raise ValueError(_("Please upload an Excel file."))

        data = base64.b64decode(self.category_import_file)
        workbook = openpyxl.load_workbook(io.BytesIO(data))
        sheet = workbook.active

        updated = 0
        skipped = 0
        max_rows = 1000
        processed = 0

        Partner = self.env['res.partner']
        Category = self.env['res.partner.category']

        # Expected headers:
        # name | google_map_url | mobile | phone | vat | company_type | category_names | customer_rank

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if processed >= max_rows:
                break
            processed += 1

            (
                name,
                google_map_url,
                mobile,
                phone,
                vat,
                company_type,
                category_names,
                customer_rank
            ) = row

            if not name:
                skipped += 1
                continue

            # 🔎 Find partner by name
            partner = Partner.search([('name', '=', name)], limit=1)

            if not partner:
                skipped += 1
                continue

            vals = {}

            # ------------------------
            # SIMPLE FIELDS
            # ------------------------
            if google_map_url:
                vals['google_map_url'] = google_map_url
            if mobile:
                vals['mobile'] = str(mobile)
            if phone:
                vals['phone'] = str(phone)
            if vat:
                vals['vat'] = vat
            if customer_rank is not None:
                vals['customer_rank'] = int(customer_rank)

            # ------------------------
            # COMPANY TYPE
            # ------------------------
            if company_type:
                vals['is_company'] = True if str(company_type).lower() == 'company' else False

            # ------------------------
            # PARTNER CATEGORIES (M2M)
            # ------------------------
            if category_names:
                category_ids = []
                for cat_name in str(category_names).split(','):
                    cat_name = cat_name.strip()
                    if not cat_name:
                        continue
                    category = Category.search([('name', '=', cat_name)], limit=1)
                    if not category:
                        category = Category.create({'name': cat_name})
                    category_ids.append(category.id)

                if category_ids:
                    vals['category_id'] = [(6, 0, category_ids)]

            if vals:
                partner.write(vals)
                updated += 1
            else:
                skipped += 1

        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Import Completed"),
                "message": _("Updated: %s\nSkipped: %s") % (updated, skipped),
                "type": "success",
            }
        }




    def action_view_due_statement(self):
        self.ensure_one()

        invoices = self.env['account.move'].sudo().search([
            ('partner_id', '=', self.id),
            ('move_type', '=', 'out_invoice'),
            ('state', '=', 'posted'),
            ('amount_residual', '>', 0),
        ])

        return {
            'type': 'ir.actions.act_window',
            'name': 'Due Invoices',
            'res_model': 'account.move',
            'domain': [('id', 'in', invoices.ids)],
            'view_mode': 'list,form',
            'target': 'current',
            'context': {
                'default_partner_id': self.id,
            },
        }
